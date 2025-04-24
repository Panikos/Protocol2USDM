import openpyxl
import json
from collections import defaultdict

# --- CONFIG ---
EXCEL_PATH = "temp/USDM_CT.xlsx"
OUTPUT_PATH = "soa_entity_mapping.json"
SHEET_ENTITIES = "DDF Entities&Attributes"
SHEET_VALUES = "DDF valid value sets"

# --- Load workbook and sheets ---
wb = openpyxl.load_workbook(EXCEL_PATH)
ws_entities = wb[SHEET_ENTITIES]
ws_values = wb[SHEET_VALUES]

# --- Parse value sets ---
value_sets = defaultdict(list)
header_row = 6
header = [cell.value for cell in ws_values[header_row]]
col_entity = header.index("Entity")
col_attr = header.index("Attribute")
col_codelist = header.index("Codelist C-code")
col_concept = header.index("Concept C-code")
col_term = header.index("Preferred Term (CDISC Submission Value)")

for row in ws_values.iter_rows(min_row=header_row+1):
    entity = row[col_entity].value
    attr = row[col_attr].value
    codelist = row[col_codelist].value
    concept = row[col_concept].value
    term = row[col_term].value
    if entity and attr and term:
        value_sets[(entity, attr)].append({
            "term": term,
            "concept_c_code": concept,
            "codelist_c_code": codelist
        })

# --- Parse entities, attributes, relationships ---
mapping = defaultdict(lambda: defaultdict(dict))
header = [cell.value for cell in ws_entities[1]]
col_entity = header.index("Entity Name")
col_role = header.index("Role")
col_ldm = header.index("Logical Data Model Name")
col_ccode = header.index("NCI C-code")
col_def = header.index("Definition")
col_has_val = header.index("Has Value List")

for row in ws_entities.iter_rows(min_row=2):
    entity = row[col_entity].value
    role = row[col_role].value
    ldm_name = row[col_ldm].value
    c_code = row[col_ccode].value
    definition = row[col_def].value
    has_val = row[col_has_val].value
    if not entity or not role:
        continue
    if role == "Entity":
        continue  # Only collect fields under entities
    # Find parent entity (previous non-empty Entity row above)
    parent_entity = entity
    # Attributes, Relationships, Complex Datatype Relationships
    field_info = {
        "name": ldm_name,
        "c_code": c_code,
        "definition": definition,
        "role": role
    }
    # Add allowed values if present
    if has_val and has_val.upper() == "Y":
        allowed = value_sets.get((parent_entity, ldm_name), [])
        if allowed:
            field_info["allowed_values"] = allowed
    if role == "Attribute":
        mapping[parent_entity]["attributes"][ldm_name] = field_info
    elif role == "Relationship":
        mapping[parent_entity]["relationships"][ldm_name] = field_info
    elif role == "Complex Datatype Relationship":
        mapping[parent_entity]["complex_datatype_relationships"][ldm_name] = field_info

# --- Write to JSON ---
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(mapping, f, indent=2, ensure_ascii=False)

print(f"[SUCCESS] Entity mapping written to {OUTPUT_PATH}")
